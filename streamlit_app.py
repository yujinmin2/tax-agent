"""
세무사 에이전트 A — 웹 버전 (RAG + 대화 저장)
Streamlit Cloud 배포용
"""

import base64
import io
import json
import uuid
from datetime import datetime, timezone, timedelta

import streamlit as st
import anthropic
from pathlib import Path
from tax_knowledge_2026 import BASE_PROMPT
from rag import LawRetriever

KST = timezone(timedelta(hours=9))

st.set_page_config(
    page_title="세무사 에이전트 A",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("🏛️ 세무사 에이전트 A")
st.caption("2026년 한국 세법 전문 AI · claude-opus-4-7 · Adaptive Thinking · RAG")


# ── RAG 인덱스 로드 (앱 수명 동안 1회) ─────────────
@st.cache_resource
def get_retriever():
    root = Path(__file__).parent
    chunks = root / "law_chunks.json"
    embeddings = root / "law_embeddings.npy"
    if not chunks.exists():
        st.error("law_chunks.json 없음. build_index.py 먼저 실행하세요.")
        st.stop()
    voyage_key = st.secrets.get("VOYAGE_API_KEY", "")
    return LawRetriever(chunks, embeddings, voyage_key)


@st.cache_resource
def get_client():
    return anthropic.Anthropic(api_key=st.secrets["ANTHROPIC_API_KEY"])


@st.cache_resource
def get_supabase():
    """Supabase 클라이언트 (SUPABASE_URL/KEY 없으면 None 반환)"""
    url = st.secrets.get("SUPABASE_URL", "")
    key = st.secrets.get("SUPABASE_KEY", "")
    if not url or not key:
        return None
    try:
        from supabase import create_client
        return create_client(url, key)
    except Exception:
        return None


retriever = get_retriever()
client = get_client()
sb = get_supabase()


# ── 파일 처리 ─────────────────────────────────────
def process_file(uploaded_file) -> dict | None:
    """업로드된 파일을 Anthropic content block으로 변환"""
    name = uploaded_file.name.lower()
    raw = uploaded_file.read()

    # PDF
    if name.endswith(".pdf"):
        b64 = base64.standard_b64encode(raw).decode()
        return {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }

    # 이미지
    if name.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
        if name.endswith(".png"):
            mime = "image/png"
        elif name.endswith((".jpg", ".jpeg")):
            mime = "image/jpeg"
        elif name.endswith(".webp"):
            mime = "image/webp"
        else:
            mime = "image/gif"
        b64 = base64.standard_b64encode(raw).decode()
        return {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}}

    # Excel
    if name.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"[시트: {sheet}]")
                for row in ws.iter_rows(values_only=True):
                    row_str = "\t".join("" if v is None else str(v) for v in row)
                    if row_str.strip():
                        lines.append(row_str)
            text = "\n".join(lines)
        except Exception as e:
            text = f"Excel 파일 파싱 오류: {e}"
        return {"type": "text", "text": f"[첨부 Excel: {uploaded_file.name}]\n{text}"}

    # CSV
    if name.endswith(".csv"):
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp949", errors="replace")
        return {"type": "text", "text": f"[첨부 CSV: {uploaded_file.name}]\n{text}"}

    return None


# ── 대화 저장 관련 ─────────────────────────────────
def _sanitize_messages(messages: list) -> list:
    """이미지/PDF 바이너리는 플레이스홀더로 교체 후 저장 (DB 용량 절약)"""
    result = []
    for msg in messages:
        content = msg["content"]
        if isinstance(content, list):
            new_blocks = []
            for blk in content:
                if isinstance(blk, dict) and blk.get("type") == "image":
                    new_blocks.append({"type": "text", "text": "📷 [이미지 첨부 — 재열람 시 재첨부 필요]"})
                elif isinstance(blk, dict) and blk.get("type") == "document":
                    new_blocks.append({"type": "text", "text": "📄 [PDF 첨부 — 재열람 시 재첨부 필요]"})
                else:
                    new_blocks.append(blk)
            result.append({"role": msg["role"], "content": new_blocks})
        else:
            result.append(msg)
    return result


def save_conversation(session_id: str, title: str, messages: list):
    """Supabase에 대화 upsert (없으면 INSERT, 있으면 UPDATE)"""
    if not sb or not messages:
        return
    try:
        sanitized = _sanitize_messages(messages)
        sb.table("conversations").upsert({
            "session_id": session_id,
            "title": title[:60],
            "messages": sanitized,
            "updated_at": datetime.now(tz=timezone.utc).isoformat(),
        }, on_conflict="session_id").execute()
    except Exception as e:
        st.toast(f"저장 오류: {e}", icon="⚠️")


def load_conversations(limit: int = 30) -> list:
    """최근 대화 목록 반환"""
    if not sb:
        return []
    try:
        res = (
            sb.table("conversations")
            .select("session_id, title, updated_at")
            .order("updated_at", desc=True)
            .limit(limit)
            .execute()
        )
        return res.data or []
    except Exception:
        return []


def load_conversation_messages(session_id: str) -> list:
    """특정 대화의 messages 반환"""
    if not sb:
        return []
    try:
        res = (
            sb.table("conversations")
            .select("messages")
            .eq("session_id", session_id)
            .single()
            .execute()
        )
        return res.data.get("messages", []) if res.data else []
    except Exception:
        return []


def _fmt_date(iso_str: str) -> str:
    """UTC ISO → 한국시간 '05/22 14:30' 형식"""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00")).astimezone(KST)
        today = datetime.now(tz=KST).date()
        if dt.date() == today:
            return f"오늘 {dt:%H:%M}"
        elif (today - dt.date()).days == 1:
            return f"어제 {dt:%H:%M}"
        else:
            return dt.strftime("%m/%d %H:%M")
    except Exception:
        return iso_str[:10]


# ── 세션 초기화 ────────────────────────────────────
def _empty_stats():
    return {"input": 0, "cache_read": 0, "cache_write": 0, "output": 0, "turns": 0}


def _new_session():
    st.session_state.session_id = str(uuid.uuid4())
    st.session_state.messages = []
    st.session_state.stats = _empty_stats()
    st.session_state.conv_title = "새 대화"


if "session_id" not in st.session_state:
    _new_session()
if "messages" not in st.session_state:
    st.session_state.messages = []
if "stats" not in st.session_state:
    st.session_state.stats = _empty_stats()
if "conv_title" not in st.session_state:
    st.session_state.conv_title = "새 대화"


# ── 사이드바 ──────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 메뉴")

    col_new, col_del = st.columns(2)
    with col_new:
        if st.button("🆕 새 대화", use_container_width=True):
            _new_session()
            st.rerun()
    with col_del:
        if st.button("🗑️ 초기화", use_container_width=True):
            _new_session()
            st.rerun()

    st.divider()

    # ── 파일 첨부 ────────────────────────────────
    st.subheader("📎 파일 첨부")
    uploaded_file = st.file_uploader(
        "PDF · 이미지 · Excel · CSV",
        type=["pdf", "png", "jpg", "jpeg", "webp", "gif", "xlsx", "xls", "csv"],
        help="세금계산서, 재무제표, 사업계획서 등을 첨부하면 AI가 함께 분석합니다.",
    )
    if uploaded_file:
        st.caption(f"✅ {uploaded_file.name} ({uploaded_file.size/1024:.1f} KB)")

    st.divider()
    st.subheader("📊 토큰 사용량")

    s = st.session_state.stats
    col1, col2 = st.columns(2)
    col1.metric("대화 횟수", s["turns"])
    col2.metric("출력 토큰", f"{s['output']:,}")
    st.caption(
        f"입력 {s['input']:,} | 캐시읽기 {s['cache_read']:,} | 캐시쓰기 {s['cache_write']:,}"
    )

    # ── 대화 기록 ────────────────────────────────
    if sb:
        st.divider()
        st.subheader("💾 대화 기록")
        conv_list = load_conversations(30)
        if conv_list:
            for conv in conv_list:
                sid = conv["session_id"]
                title = conv.get("title", "새 대화")
                date_str = _fmt_date(conv.get("updated_at", ""))
                label = f"{date_str}  {title}"
                is_current = sid == st.session_state.session_id
                btn_type = "primary" if is_current else "secondary"
                if st.button(label, key=f"conv_{sid}", use_container_width=True, type=btn_type):
                    if not is_current:
                        msgs = load_conversation_messages(sid)
                        st.session_state.session_id = sid
                        st.session_state.messages = msgs
                        st.session_state.stats = _empty_stats()
                        st.session_state.conv_title = title
                        st.rerun()
        else:
            st.caption("저장된 대화가 없습니다.")

    st.divider()
    st.markdown(
        """
**주요 기능**
- 소득세 · 법인세 · 부가가치세
- 상속·증여세 · 종합부동산세
- 양도소득세 · 원천징수
- 세무신고 절차 · 절세 전략
        """
    )


# ── 대화 내역 출력 ─────────────────────────────────
def _extract_text(content) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        return " ".join(
            b.get("text", "") for b in content
            if isinstance(b, dict) and b.get("type") == "text"
        )
    return ""


for msg in st.session_state.messages:
    avatar = "👤" if msg["role"] == "user" else "🏛️"
    with st.chat_message(msg["role"], avatar=avatar):
        content = msg["content"]
        if isinstance(content, str):
            st.markdown(content)
        elif isinstance(content, list):
            for block in content:
                if isinstance(block, dict):
                    if block.get("type") == "text":
                        st.markdown(block["text"])
                    elif block.get("type") == "image":
                        img_data = base64.b64decode(block["source"]["data"])
                        st.image(img_data)
                    elif block.get("type") == "document":
                        st.caption("📄 PDF 문서 첨부됨")
                else:
                    st.markdown(str(block))


# ── 입력 처리 ──────────────────────────────────────
if prompt := st.chat_input("세금 관련 질문을 입력하세요..."):
    # 파일 첨부 여부에 따라 content 구성
    if uploaded_file is not None:
        uploaded_file.seek(0)
        file_block = process_file(uploaded_file)
        if file_block:
            user_content = [file_block, {"type": "text", "text": prompt}]
        else:
            st.warning(f"지원하지 않는 파일 형식입니다: {uploaded_file.name}")
            user_content = prompt
    else:
        user_content = prompt

    # 대화 제목: 첫 메시지로 설정
    if not st.session_state.messages:
        st.session_state.conv_title = (_extract_text(user_content) or "새 대화")[:40]

    st.session_state.messages.append({"role": "user", "content": user_content})

    with st.chat_message("user", avatar="👤"):
        if uploaded_file and isinstance(user_content, list):
            fname = uploaded_file.name.lower()
            if fname.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
                for blk in user_content:
                    if isinstance(blk, dict) and blk.get("type") == "image":
                        st.image(base64.b64decode(blk["source"]["data"]))
            elif fname.endswith(".pdf"):
                st.caption(f"📄 {uploaded_file.name}")
            elif fname.endswith((".xlsx", ".xls", ".csv")):
                st.caption(f"📊 {uploaded_file.name}")
        st.markdown(prompt)

    with st.chat_message("assistant", avatar="🏛️"):
        status_placeholder = st.empty()
        response_placeholder = st.empty()
        status_placeholder.markdown("*⏳ 관련 조문 검색 중...*")

        # ── RAG: 관련 조문 검색 ─────────────────────
        recent_ctx = " ".join(
            _extract_text(m["content"]) for m in st.session_state.messages[-6:]
        )
        retrieved = retriever.retrieve(recent_ctx, top_k=25)

        # 시스템 프롬프트 구성
        system_blocks = [
            {
                "type": "text",
                "text": BASE_PROMPT,
                "cache_control": {"type": "ephemeral"},
            }
        ]
        if retrieved:
            system_blocks.append({
                "type": "text",
                "text": (
                    "\n\n══════════════════════════════\n"
                    "【이번 질문과 관련된 실제 법령 조문 (law.go.kr 2026)】\n"
                    "아래 조문을 우선 인용하여 답변하세요.\n"
                    "══════════════════════════════\n\n"
                    + retrieved
                ),
            })

        status_placeholder.markdown("*⏳ 분석 중...*")
        full_response = ""
        started = False

        with client.messages.stream(
            model="claude-opus-4-7",
            max_tokens=8192,
            thinking={"type": "adaptive"},
            system=system_blocks,
            messages=st.session_state.messages,
        ) as stream:
            for text in stream.text_stream:
                if not started:
                    started = True
                    status_placeholder.empty()
                full_response += text
                response_placeholder.markdown(full_response + "▌")

            final_msg = stream.get_final_message()

        response_placeholder.markdown(full_response)

        # 사용량 집계
        usage = final_msg.usage
        s = st.session_state.stats
        s["input"] += getattr(usage, "input_tokens", 0) or 0
        s["cache_read"] += getattr(usage, "cache_read_input_tokens", 0) or 0
        s["cache_write"] += getattr(usage, "cache_creation_input_tokens", 0) or 0
        s["output"] += getattr(usage, "output_tokens", 0) or 0
        s["turns"] += 1

    st.session_state.messages.append({"role": "assistant", "content": full_response})

    # ── 대화 자동 저장 ───────────────────────────
    save_conversation(
        st.session_state.session_id,
        st.session_state.conv_title,
        st.session_state.messages,
    )
